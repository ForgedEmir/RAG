import asyncio
import io
import logging
import mimetypes
import os
import re
import tarfile
import zipfile
from fastapi import APIRouter, Depends, Request, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
from src.api.auth import get_current_user, get_tenant_id, require_monitoring
from src.api.limiter import limiter
from src.ingestion.run import index_data, SUPPORTED_EXTENSIONS
from src.monitoring.tracker import track
from src.security.validator import valider_entree

logger = logging.getLogger(__name__)
admin_router = APIRouter()

MAX_UPLOAD_SIZE = int(os.getenv("MAX_UPLOAD_SIZE_KB", "500")) * 1024
MAX_ARCHIVE_SIZE = int(os.getenv("MAX_ARCHIVE_SIZE_MB", "50")) * 1024 * 1024
ALLOWED_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".xml", ".xlsx", ".pdf", ".docx", ".doc"}
ARCHIVE_EXTENSIONS = {".zip", ".tar.gz", ".tar.bz2", ".tar.xz"}
MAX_ARCHIVE_FILES = 50
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 100 * 1024 * 1024  # 100 MB
DATA_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "data", "sample"))
_SAFE_FILENAME_RE = re.compile(r"^[A-Za-z0-9._ -]{1,50}$")

_EXT_TO_MIME: dict[str, str] = {
    ".pdf":  "application/pdf",
    ".txt":  "text/plain",
    ".md":   "text/markdown",
    ".csv":  "text/csv",
    ".json": "application/json",
    ".xml":  "application/xml",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc":  "application/msword",
}


def _sanitize_filename(raw_name: str) -> str:
    """Return a safe basename or an empty string when invalid."""
    name = os.path.basename((raw_name or "").replace("\x00", "")).strip()
    if not name or name in {".", ".."}:
        return ""
    if any(sep in name for sep in ("/", "\\")):
        return ""
    if ".." in name:
        return ""
    if not _SAFE_FILENAME_RE.match(name):
        return ""
    return name


def _is_archive(filename: str) -> bool:
    name = filename.lower()
    return any(name.endswith(ext) for ext in ARCHIVE_EXTENSIONS)


def _extract_archive(content: bytes, filename: str, target_dir: str) -> list[str]:
    """Extract archive safely into target_dir. Returns list of filenames written to disk.

    Enforces: zip-slip protection, zip-bomb limit, file count limit, extension filtering.
    """
    name = filename.lower()
    real_target = os.path.realpath(target_dir)
    extracted: list[str] = []

    def _safe_write(data: bytes, basename: str) -> str | None:
        safe = _sanitize_filename(basename)
        if not safe:
            return None
        if os.path.splitext(safe)[1].lower() not in ALLOWED_EXTENSIONS:
            return None
        dest = os.path.join(target_dir, safe)
        # Zip-slip guard: resolved destination must stay inside target_dir
        if not os.path.realpath(dest).startswith(real_target + os.sep):
            logger.warning("[ARCHIVE] Zip-slip bloqué : '%s'", basename)
            return None
        with open(dest, "wb") as f:
            f.write(data)
        return safe

    if name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            infos = [i for i in zf.infolist() if not i.is_dir()]
            total = sum(i.file_size for i in infos)
            if total > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                raise ValueError(
                    f"Archive trop volumineuse décompressée "
                    f"({total // 1024 // 1024} Mo > {MAX_ARCHIVE_UNCOMPRESSED_BYTES // 1024 // 1024} Mo max)"
                )
            if len(infos) > MAX_ARCHIVE_FILES:
                raise ValueError(f"Trop de fichiers dans l'archive ({len(infos)} > {MAX_ARCHIVE_FILES} max)")
            for info in infos:
                if os.path.isabs(info.filename) or ".." in info.filename.replace("\\", "/").split("/"):
                    logger.warning("[ARCHIVE] Zip-slip bloqué : '%s'", info.filename)
                    continue
                with zf.open(info) as src:
                    saved = _safe_write(src.read(), os.path.basename(info.filename))
                    if saved:
                        extracted.append(saved)
    else:
        mode = (
            "r:gz" if name.endswith(".tar.gz") else
            "r:bz2" if name.endswith(".tar.bz2") else
            "r:xz" if name.endswith(".tar.xz") else
            "r:"
        )
        with tarfile.open(fileobj=io.BytesIO(content), mode=mode) as tf:
            members = [m for m in tf.getmembers() if m.isfile()]
            total = sum(m.size for m in members)
            if total > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                raise ValueError(
                    f"Archive trop volumineuse décompressée "
                    f"({total // 1024 // 1024} Mo > {MAX_ARCHIVE_UNCOMPRESSED_BYTES // 1024 // 1024} Mo max)"
                )
            if len(members) > MAX_ARCHIVE_FILES:
                raise ValueError(f"Trop de fichiers dans l'archive ({len(members)} > {MAX_ARCHIVE_FILES} max)")
            for member in members:
                fobj = tf.extractfile(member)
                if fobj:
                    saved = _safe_write(fobj.read(), os.path.basename(member.name))
                    if saved:
                        extracted.append(saved)

    return extracted


def _count_points_for_filename(filename: str) -> int | None:
    try:
        from qdrant_client.models import Filter, FieldCondition, MatchValue
        from src.ingestion.vector_store import get_store
        store = get_store(force_reindex=False)
        result = store.client.count(
            collection_name="knowledge",
            count_filter=Filter(
                should=[FieldCondition(key="metadata.fichier", match=MatchValue(value=filename))]
            ),
            exact=True,
        )
        return int(getattr(result, "count", 0))
    except Exception as e:
        logger.warning(f"Vérification Qdrant impossible pour '{filename}': {e}")
        return None


def _read_sample(content: bytes, ext: str, safe_name: str) -> str:
    """Extrait un échantillon de texte pour la validation sécurité."""
    text = content.decode("utf-8", errors="ignore") if ext in {".txt", ".md", ".csv", ".json", ".xml"} else ""
    lines = text.splitlines()
    if len(lines) > 40:
        mid = len(lines) // 2
        return "\n".join(lines[:20] + lines[max(0, mid - 5):mid + 5] + lines[-20:])
    if not text and ext in {".pdf", ".xlsx", ".docx", ".doc"}:
        return f"BINARY_FILE:{ext}:{safe_name}:{len(content)}"
    return text


# ── Endpoint /api/upload — accessible à tous les users authentifiés ───────────

@admin_router.post("/api/upload")
@limiter.limit("100/hour")
async def tenant_upload(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user),
):
    """Upload multi-tenant : valide → Supabase Storage → ingestion Qdrant avec tenant_id."""
    tenant_id = await get_tenant_id(user_id)

    safe_name = _sanitize_filename(file.filename or "")
    if not safe_name:
        return JSONResponse({"error": "Nom de fichier invalide."}, status_code=400)

    is_arch = _is_archive(safe_name)
    ext = os.path.splitext(safe_name)[1].lower()
    if not is_arch and ext not in ALLOWED_EXTENSIONS:
        all_formats = sorted(ALLOWED_EXTENSIONS | ARCHIVE_EXTENSIONS)
        return JSONResponse(
            {"error": f"Extension non supportée. Formats : {', '.join(all_formats)}"},
            status_code=400,
        )

    content = await file.read()
    if not content:
        return JSONResponse({"error": "Fichier vide."}, status_code=400)

    size_limit = MAX_ARCHIVE_SIZE if is_arch else MAX_UPLOAD_SIZE
    if len(content) > size_limit:
        size_kb = len(content) // 1024
        return JSONResponse(
            {"error": f"Fichier trop volumineux ({size_kb} Ko). Max : {size_limit // 1024} Ko."},
            status_code=400,
        )

    tenant_data_dir = os.path.join(DATA_DIR, tenant_id)
    os.makedirs(tenant_data_dir, exist_ok=True)

    # ── Archive path ─────────────────────────────────────────────────────────
    if is_arch:
        try:
            extracted_names = _extract_archive(content, safe_name, tenant_data_dir)
        except ValueError as e:
            return JSONResponse({"error": str(e)}, status_code=400)

        if not extracted_names:
            return JSONResponse(
                {"error": "Aucun fichier valide trouvé dans l'archive (formats acceptés : "
                          f"{', '.join(sorted(ALLOWED_EXTENSIONS))})."},
                status_code=400,
            )

        storage_ok_count = 0
        try:
            from src.monitoring.tracker import _get_client
            supa = await _get_client()
            if supa:
                for fname in extracted_names:
                    fpath = os.path.join(tenant_data_dir, fname)
                    with open(fpath, "rb") as fh:
                        fdata = fh.read()
                    try:
                        fext = os.path.splitext(fname)[1].lower()
                        await supa.storage.from_("tenant-docs").upload(
                            path=f"{tenant_id}/{fname}",
                            file=fdata,
                            file_options={"content-type": _EXT_TO_MIME.get(fext, "application/octet-stream"), "upsert": "true"},
                        )
                        storage_ok_count += 1
                    except Exception as e:
                        logger.warning("[UPLOAD] Supabase Storage échoué pour '%s/%s': %s", tenant_id, fname, e)
        except Exception as e:
            logger.warning("[UPLOAD] Supabase Storage archive échoué: %s", e)

        try:
            from src.ingestion.run import index_data as _index
            await asyncio.to_thread(_index, False, tenant_id)
            logger.info("[UPLOAD] Ingestion archive OK — tenant=%s fichiers=%s", tenant_id, extracted_names)
        except Exception as e:
            logger.error("[UPLOAD] Ingestion archive échouée — tenant=%s : %s", tenant_id, e)
        await track("upload", detail=f"tenant={tenant_id} | archive={safe_name} | {len(extracted_names)} fichiers")
        return {
            "message": f"{len(extracted_names)} fichier(s) extrait(s) depuis '{safe_name}'. Ingestion en cours.",
            "files": extracted_names,
            "count": len(extracted_names),
            "tenant_id": tenant_id,
            "storage_ok": storage_ok_count,
        }

    # ── Regular file path (unchanged) ────────────────────────────────────────
    sample = _read_sample(content, ext, safe_name)
    validation = await valider_entree(sample)
    if not validation["valid"]:
        logger.warning(f"[UPLOAD] Bloqué [{validation['type']}] tenant={tenant_id} fichier='{safe_name}'")
        await track("upload_blocked", detail=f"tenant={tenant_id} | {safe_name} | {validation['type']}")
        return JSONResponse({"error": "Contenu suspect. Upload refusé."}, status_code=400)

    # ── Supabase Storage (bucket "tenant-docs", path "tenant_id/filename") ──
    storage_path = f"{tenant_id}/{safe_name}"
    storage_ok = False
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if supa:
            await supa.storage.from_("tenant-docs").upload(
                path=storage_path,
                file=content,
                file_options={
                    "content-type": _EXT_TO_MIME.get(ext, file.content_type or "application/octet-stream"),
                    "upsert": "true",
                },
            )
            storage_ok = True
            logger.info(f"[UPLOAD] Supabase Storage OK : {storage_path}")
    except Exception as e:
        logger.warning(f"[UPLOAD] Supabase Storage échoué pour '{storage_path}': {e}")

    # ── Écriture locale dans data/sample/<tenant_id>/ pour le pipeline ──────
    destination = os.path.join(tenant_data_dir, safe_name)
    existed_before = os.path.exists(destination)
    with open(destination, "wb") as f:
        f.write(content)

    # ── Ingestion synchrone — le 200 est renvoyé après indexation ───────────
    try:
        from src.ingestion.run import index_data as _index
        await asyncio.to_thread(_index, False, tenant_id)
        logger.info(f"[UPLOAD] Ingestion OK — tenant={tenant_id} fichier={safe_name}")
    except Exception as e:
        logger.error(f"[UPLOAD] Ingestion échouée — tenant={tenant_id} : {e}")

    action = "replace" if existed_before else "upload"
    await track(action, detail=f"tenant={tenant_id} | {safe_name} | {len(content) // 1024} Ko")
    return {
        "message": f"'{safe_name}' {'remplacé' if existed_before else 'uploadé'}. Ingestion en cours.",
        "filename": safe_name,
        "tenant_id": tenant_id,
        "storage_path": storage_path,
        "storage_ok": storage_ok,
    }


# ── Routes admin ──────────────────────────────────────────────────────────────

@admin_router.get("/api/admin/sources")
async def admin_sources(request: Request):
    require_monitoring(request)
    from src.ingestion.run import list_current_files
    files = list_current_files()
    return {"files": sorted(files.keys()), "total": len(files)}


@admin_router.get("/api/sources")
async def user_sources(user_id: str = Depends(get_current_user)):
    tenant_id = await get_tenant_id(user_id)
    try:
        from src.ingestion.vector_store import _get_client, _COLLECTION_NAME
        client = _get_client()
        info = client.get_collection(_COLLECTION_NAME)
        total_pts = info.points_count or 0
        logger.info("[SOURCES] collection=%s total_points=%d tenant=%s", _COLLECTION_NAME, total_pts, tenant_id)
        filenames: set[str] = set()
        offset = None
        filt = None
        while True:
            results, next_offset = client.scroll(
                collection_name=_COLLECTION_NAME,
                scroll_filter=filt,
                limit=256,
                offset=offset,
                with_payload=True,
                with_vectors=False,
            )
            for point in results:
                payload = point.payload or {}
                meta = payload.get("metadata") or {}
                fichier = meta.get("fichier") or payload.get("fichier")
                if fichier:
                    filenames.add(os.path.basename(fichier))
            if next_offset is None:
                break
            offset = next_offset
        logger.info("[SOURCES] found %d unique files", len(filenames))
        files = sorted(filenames)
        return {"files": files, "total": len(files)}
    except Exception as e:
        logger.error("[SOURCES] Qdrant error: %s", e, exc_info=True)
        return {"files": [], "total": 0, "error": str(e)}


@admin_router.post("/api/admin/upload")
@limiter.limit("20/hour")
async def admin_upload(request: Request, file: UploadFile = File(...)):
    require_monitoring(request)

    safe_name = _sanitize_filename(file.filename or "")
    if not safe_name:
        return JSONResponse({"error": "Nom de fichier invalide."}, status_code=400)

    is_arch = _is_archive(safe_name)
    ext = os.path.splitext(safe_name)[1].lower()
    if not is_arch and ext not in ALLOWED_EXTENSIONS:
        return JSONResponse({"error": f"Extension non supportée. Formats : {', '.join(ALLOWED_EXTENSIONS)}"}, status_code=400)

    content = await file.read()
    if not content:
        return JSONResponse({"error": "Fichier vide."}, status_code=400)

    size_limit = MAX_ARCHIVE_SIZE if is_arch else MAX_UPLOAD_SIZE
    if len(content) > size_limit:
        size_kb, limit_kb = len(content) // 1024, size_limit // 1024
        return JSONResponse({"error": f"Fichier trop volumineux ({size_kb} Ko). Max : {limit_kb} Ko."}, status_code=400)

    os.makedirs(DATA_DIR, exist_ok=True)

    # ── Archive path ─────────────────────────────────────────────────────────
    if is_arch:
        try:
            extracted_names = _extract_archive(content, safe_name, DATA_DIR)
        except ValueError as e:
            return JSONResponse({"error": str(e)}, status_code=400)

        if not extracted_names:
            return JSONResponse(
                {"error": f"Aucun fichier valide trouvé dans l'archive (formats acceptés : {', '.join(sorted(ALLOWED_EXTENSIONS))})."},
                status_code=400,
            )

        changed = None
        reindex_warning = None
        try:
            changed = bool(await asyncio.to_thread(index_data, force_reindex=False))
        except Exception as e:
            reindex_warning = str(e)
            logger.warning("Admin archive réindexation échouée: %s", e)

        await track("upload", detail=f"archive={safe_name} | {len(extracted_names)} fichiers")
        payload = {
            "message": f"{len(extracted_names)} fichier(s) extrait(s) depuis '{safe_name}' et indexé(s).",
            "files": extracted_names,
            "count": len(extracted_names),
            "ingestion": {"changed": changed},
        }
        if reindex_warning:
            payload["warning"] = "Fichiers extraits, mais la réindexation automatique a échoué."
        return payload

    # ── Regular file path (unchanged) ────────────────────────────────────────
    try:
        text = content.decode("utf-8", errors="ignore") if ext in {".txt", ".md", ".csv", ".json", ".xml"} else ""
    except Exception:
        return JSONResponse({"error": "Encodage invalide."}, status_code=400)

    # Security check on a representative sample to avoid high latency on large files
    lines = text.splitlines()
    if len(lines) > 50:
        mid = len(lines) // 2
        sample = "\n".join(lines[:20] + lines[max(0, mid - 5):mid + 5] + lines[-20:])
    else:
        sample = text

    # Binary formats: shallow marker for validator path, keep low latency.
    if not sample and ext in {".pdf", ".xlsx"}:
        sample = f"BINARY_FILE:{ext}:{safe_name}:{len(content)}"

    validation = await valider_entree(sample)
    if not validation["valid"]:
        logger.warning(f"Upload blocked [{validation['type']}] — '{safe_name}'")
        await track("upload_blocked", detail=f"{safe_name} | {validation['type']}")
        return JSONResponse({"error": "Contenu suspect. Upload refusé."}, status_code=400)

    destination = os.path.join(DATA_DIR, safe_name)
    existed_before = os.path.exists(destination)

    with open(destination, "wb") as out:
        out.write(content)

    reindex_warning = None
    changed = None
    try:
        # index_data est synchrone
        changed = bool(await asyncio.to_thread(index_data, force_reindex=False))
    except Exception as e:
        reindex_warning = str(e)
        logger.warning(f"Upload réindexation automatique échouée: {e}")

    qdrant_points = _count_points_for_filename(safe_name)

    if existed_before:
        await track("replace", detail=f"{safe_name} | {len(content)//1024} Ko (upsert)")
        logger.info(f"Fichier remplacé (upsert upload) : {safe_name}")
        payload = {"message": f"'{safe_name}' remplacé et indexé.", "filename": safe_name}
    else:
        await track("upload", detail=f"{safe_name} | {len(content)//1024} Ko")
        logger.info(f"Fichier uploadé : {safe_name}")
        payload = {"message": f"'{safe_name}' uploadé et indexé.", "filename": safe_name}

    payload["ingestion"] = {
        "changed": changed,
        "qdrant_points": qdrant_points,
        "verified": (qdrant_points is not None and qdrant_points > 0),
        "summary": (
            f"Ingestion {'mise à jour' if changed else 'sans changement détecté'}; "
            + (f"Qdrant={qdrant_points} chunk(s)." if qdrant_points is not None else "Qdrant non vérifiable.")
        ),
    }
    if reindex_warning:
        payload["warning"] = "Fichier uploadé, mais la réindexation automatique a échoué."
    return payload


@admin_router.post("/api/admin/invite")
@limiter.limit("20/hour")
async def invite_client(request: Request):
    """Invite un client par email via Supabase Admin API (magic link automatique)."""
    require_monitoring(request)
    body = await request.json()
    email = (body or {}).get("email", "").strip().lower()

    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return JSONResponse({"error": "Adresse email invalide."}, status_code=400)

    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        app_url = os.getenv("APP_URL", "http://localhost:8000")
        redirect_url = f"{app_url}/auth/callback"
        try:
            res = await supa.auth.admin.invite_user_by_email(
                email,
                options={"redirect_to": redirect_url},
            )
        except TypeError:
            # Certaines versions gotrue-py n'acceptent pas options= comme kwarg
            res = await supa.auth.admin.invite_user_by_email(email)
        user = getattr(res, "user", None)
        user_id = str(user.id) if user and hasattr(user, "id") else None
        logger.info(f"[INVITE] Client invité : {email} (user_id={user_id})")
        await track("invite", detail=f"email={email}")
        return {"message": f"Invitation envoyée à {email}.", "user_id": user_id}
    except Exception as e:
        err = str(e)
        if "already registered" in err.lower() or "already been registered" in err.lower():
            return JSONResponse({"error": "Cet email est déjà enregistré."}, status_code=409)
        logger.error(f"[INVITE] Échec invitation {email}: {e}")
        return JSONResponse({"error": f"Échec de l'invitation : {err}"}, status_code=500)


@admin_router.get("/api/admin/clients")
async def list_clients(request: Request):
    """Liste les utilisateurs Supabase (clients)."""
    require_monitoring(request)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        res = await supa.auth.admin.list_users()
        # list_users() peut retourner une liste ou un objet paginé selon la version gotrue-py
        if isinstance(res, list):
            raw = res
        elif hasattr(res, "users"):
            raw = res.users or []
        else:
            raw = list(res) if res else []

        def _str(val):
            if val is None:
                return None
            if isinstance(val, str):
                return val
            if hasattr(val, "isoformat"):
                return val.isoformat()
            return str(val)

        users = [
            {
                "id": _str(u.id),
                "email": u.email or "",
                "created_at": _str(u.created_at),
                "last_sign_in_at": _str(u.last_sign_in_at),
                "confirmed": bool(u.email_confirmed_at),
            }
            for u in raw
        ]
        return {"clients": users, "total": len(users)}
    except Exception as e:
        logger.error(f"[CLIENTS] Erreur liste clients: {e}", exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.delete("/api/admin/clients/{user_id}")
async def delete_client(user_id: str, request: Request):
    """Supprime un client Supabase par son user_id."""
    require_monitoring(request)
    if not user_id or len(user_id) != 36:
        return JSONResponse({"error": "user_id invalide."}, status_code=400)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        await supa.auth.admin.delete_user(user_id)
        logger.info(f"[CLIENTS] Utilisateur supprimé : {user_id}")
        await track("client_delete", detail=f"user_id={user_id}")
        return {"message": "Utilisateur supprimé.", "user_id": user_id}
    except Exception as e:
        logger.error(f"[CLIENTS] Erreur suppression {user_id}: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.post("/api/team/invite")
@limiter.limit("10/hour")
async def team_invite(request: Request, user_id: str = Depends(get_current_user)):
    """Invite a collaborator to the caller's tenant."""
    body = await request.json()
    email = (body or {}).get("email", "").strip().lower()
    role  = (body or {}).get("role", "member").strip().lower()

    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return JSONResponse({"error": "Adresse email invalide."}, status_code=400)
    if role not in ("admin", "member", "viewer"):
        return JSONResponse({"error": "Rôle invalide (admin, member, viewer)."}, status_code=400)

    tenant_id = await get_tenant_id(user_id)

    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)

        # Check if already a member
        existing = await supa.table("user_roles").select("id, user_id").eq("tenant_id", tenant_id).execute()
        member_emails = set()
        if existing.data:
            # Fetch emails for existing member user_ids
            ids = [r["user_id"] for r in existing.data]
            for uid in ids:
                try:
                    u = await supa.auth.admin.get_user_by_id(uid)
                    if u and u.user and u.user.email:
                        member_emails.add(u.user.email.lower())
                except Exception:
                    pass
        if email in member_emails:
            return JSONResponse({"error": "Cet utilisateur est déjà membre."}, status_code=409)

        # Check pending invitation
        pending = await supa.table("invitations").select("id").eq("tenant_id", tenant_id).eq("email", email).is_("accepted_at", "null").execute()
        if pending.data:
            return JSONResponse({"error": "Une invitation est déjà en attente pour cet email."}, status_code=409)

        app_url = os.getenv("APP_URL", "http://localhost:8000")
        redirect_url = f"{app_url}/auth/callback"
        try:
            res = await supa.auth.admin.invite_user_by_email(email, options={"redirect_to": redirect_url})
        except TypeError:
            res = await supa.auth.admin.invite_user_by_email(email)

        invited_user_id = str(res.user.id) if res and res.user else None

        await supa.table("invitations").insert({
            "email": email,
            "role": role,
            "tenant_id": tenant_id,
            "invited_by": user_id,
        }).execute()

        await track("team_invite", detail=f"tenant={tenant_id} email={email} role={role}")
        return {"message": f"Invitation envoyée à {email}.", "email": email, "role": role}
    except Exception as e:
        err = str(e)
        if "already registered" in err.lower() or "already been registered" in err.lower():
            # User exists — just insert invitation row; they'll get a notification
            try:
                from src.monitoring.tracker import _get_client as _gc2
                supa2 = await _gc2()
                if supa2:
                    await supa2.table("invitations").insert({
                        "email": email, "role": role,
                        "tenant_id": tenant_id, "invited_by": user_id,
                    }).execute()
                return {"message": f"Invitation enregistrée pour {email}.", "email": email, "role": role}
            except Exception:
                pass
        logger.error(f"[TEAM_INVITE] Erreur: {e}")
        return JSONResponse({"error": f"Échec : {err}"}, status_code=500)


@admin_router.post("/api/team/join")
async def team_join(user_id: str = Depends(get_current_user)):
    """Auto-accepte les invitations en attente pour l'email du user connecté."""
    try:
        from src.monitoring.tracker import _get_client
        from datetime import datetime, timezone
        supa = await _get_client()
        if not supa:
            return {"joined": False}

        # Récupère l'email du user depuis Supabase Auth
        u = await supa.auth.admin.get_user_by_id(user_id)
        email = (u.user.email or "").lower() if u and u.user else ""
        if not email:
            return {"joined": False}

        # Cherche les invitations en attente
        inv_res = await supa.table("invitations").select("*").eq("email", email).is_("accepted_at", "null").execute()
        if not inv_res.data:
            return {"joined": False}

        joined = []
        now_iso = datetime.now(timezone.utc).isoformat()
        for inv in inv_res.data:
            tenant = inv["tenant_id"]
            # Vérifie si déjà membre
            existing = await supa.table("user_roles").select("id").eq("user_id", user_id).eq("tenant_id", tenant).execute()
            if not existing.data:
                await supa.table("user_roles").insert({
                    "user_id": user_id,
                    "tenant_id": tenant,
                    "role": inv["role"],
                    "invited_by": inv["invited_by"],
                }).execute()
            # Marque l'invitation comme acceptée
            await supa.table("invitations").update({"accepted_at": now_iso}).eq("id", inv["id"]).execute()
            joined.append({"tenant_id": tenant, "role": inv["role"]})

        await track("team_join", detail=f"user={user_id} tenants={[j['tenant_id'] for j in joined]}")
        return {"joined": bool(joined), "tenants": joined}
    except Exception as e:
        logger.error(f"[TEAM_JOIN] Erreur: {e}")
        return {"joined": False}


@admin_router.get("/api/team/members")
async def team_members(user_id: str = Depends(get_current_user)):
    """List members of the caller's tenant."""
    tenant_id = await get_tenant_id(user_id)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)

        # S'assure que le propriétaire du tenant a bien une entrée owner
        owner_check = await supa.table("user_roles").select("id").eq("tenant_id", tenant_id).eq("user_id", tenant_id).execute()
        if not owner_check.data:
            await supa.table("user_roles").insert({
                "user_id": tenant_id,
                "tenant_id": tenant_id,
                "role": "owner",
            }).execute()

        roles_res = await supa.table("user_roles").select("user_id, role, created_at").eq("tenant_id", tenant_id).execute()
        members = []
        for row in (roles_res.data or []):
            try:
                u = await supa.auth.admin.get_user_by_id(row["user_id"])
                email = u.user.email if u and u.user else ""
            except Exception:
                email = ""
            members.append({
                "user_id": row["user_id"],
                "email": email,
                "role": row["role"],
                "created_at": str(row["created_at"]),
                "is_me": row["user_id"] == user_id,
            })

        invites_res = await supa.table("invitations").select("email, role, created_at, accepted_at, expires_at").eq("tenant_id", tenant_id).execute()
        pending = [
            {"email": r["email"], "role": r["role"], "status": "pending", "created_at": str(r["created_at"]), "expires_at": str(r["expires_at"])}
            for r in (invites_res.data or []) if not r.get("accepted_at")
        ]
        return {"members": members, "pending_invitations": pending, "tenant_id": tenant_id}
    except Exception as e:
        logger.error(f"[TEAM_MEMBERS] Erreur: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.delete("/api/team/members/{member_id}")
async def team_remove_member(member_id: str, user_id: str = Depends(get_current_user)):
    """Remove a member from the caller's tenant."""
    if not member_id or len(member_id) > 64:
        return JSONResponse({"error": "member_id invalide."}, status_code=400)
    tenant_id = await get_tenant_id(user_id)
    if member_id == user_id:
        return JSONResponse({"error": "Vous ne pouvez pas vous retirer vous-même."}, status_code=400)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        await supa.table("user_roles").delete().eq("tenant_id", tenant_id).eq("user_id", member_id).execute()
        await track("team_remove", detail=f"tenant={tenant_id} removed={member_id}")
        return {"ok": True}
    except Exception as e:
        logger.error(f"[TEAM_REMOVE] Erreur: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.delete("/api/team/invitations/{email}")
async def team_cancel_invitation(email: str, user_id: str = Depends(get_current_user)):
    """Cancel a pending invitation."""
    tenant_id = await get_tenant_id(user_id)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        await supa.table("invitations").delete().eq("tenant_id", tenant_id).eq("email", email).is_("accepted_at", "null").execute()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.get("/api/file/{filename:path}")
async def serve_file(filename: str, request: Request, user_id: str = Depends(get_current_user)):
    """Serve a file from tenant storage for in-browser preview."""
    safe_name = _sanitize_filename(os.path.basename(filename))
    if not safe_name:
        return JSONResponse({"error": "Nom de fichier invalide."}, status_code=400)

    tenant_id = await get_tenant_id(user_id)
    mime, _ = mimetypes.guess_type(safe_name)
    mime = mime or "application/octet-stream"
    disp = "inline" if (mime.startswith("text/") or mime == "application/pdf") else "attachment"
    headers = {"Content-Disposition": f'{disp}; filename="{safe_name}"'}

    # 1. Local file (fast path)
    for path in [os.path.join(DATA_DIR, tenant_id, safe_name), os.path.join(DATA_DIR, safe_name)]:
        if os.path.isfile(path):
            return FileResponse(path, media_type=mime, headers=headers)

    # 2. Supabase Storage fallback — re-download and stream
    try:
        from src.monitoring.tracker import _get_client
        from fastapi.responses import Response as FastAPIResponse
        supa = await _get_client()
        if supa:
            storage_path = f"{tenant_id}/{safe_name}"
            data = await supa.storage.from_("tenant-docs").download(storage_path)
            if data:
                # Save locally for next time
                local_path = os.path.join(DATA_DIR, tenant_id, safe_name)
                os.makedirs(os.path.dirname(local_path), exist_ok=True)
                with open(local_path, "wb") as fh:
                    fh.write(data)
                return FastAPIResponse(content=data, media_type=mime, headers=headers)
    except Exception as e:
        logger.warning(f"[FILE] Supabase Storage fallback échoué pour '{safe_name}': {e}")

    return JSONResponse({"error": "Fichier introuvable."}, status_code=404)


@admin_router.get("/api/file-text/{filename:path}")
async def serve_file_text(filename: str, request: Request, user_id: str = Depends(get_current_user)):
    """Return the extracted plain text of a file (PDF → pypdf, others → raw read)."""
    safe_name = _sanitize_filename(os.path.basename(filename))
    if not safe_name:
        return JSONResponse({"error": "Nom de fichier invalide."}, status_code=400)

    tenant_id = await get_tenant_id(user_id)
    candidates = [
        os.path.join(DATA_DIR, tenant_id, safe_name),
        os.path.join(DATA_DIR, safe_name),
    ]
    path = next((p for p in candidates if os.path.isfile(p)), None)
    if not path:
        return JSONResponse({"error": "Fichier introuvable."}, status_code=404)

    try:
        ext = os.path.splitext(safe_name)[1].lower()
        if ext == ".pdf":
            from src.ingestion.parser import _parse_pdf_pypdf
            text = _parse_pdf_pypdf(path) or ""
        elif ext in (".docx", ".doc"):
            from src.ingestion.parser import _parse_docx
            text = _parse_docx(path) or ""
        elif ext == ".xlsx":
            from src.ingestion.parser import _xlsx_to_text
            text = _xlsx_to_text(path) or ""
        else:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read()
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(text)
    except Exception as e:
        logger.error(f"file-text error for {safe_name}: {e}")
        return JSONResponse({"error": "Impossible d'extraire le texte."}, status_code=500)


@admin_router.get("/api/file-xlsx/{filename:path}")
async def serve_file_xlsx(filename: str, request: Request, user_id: str = Depends(get_current_user)):
    """Return Excel file as structured JSON: [{sheet, headers, rows}]."""
    safe_name = _sanitize_filename(os.path.basename(filename))
    if not safe_name or not safe_name.lower().endswith(".xlsx"):
        return JSONResponse({"error": "Fichier invalide."}, status_code=400)
    tenant_id = await get_tenant_id(user_id)
    candidates = [os.path.join(DATA_DIR, tenant_id, safe_name), os.path.join(DATA_DIR, safe_name)]
    path = next((p for p in candidates if os.path.isfile(p)), None)
    if not path:
        return JSONResponse({"error": "Fichier introuvable."}, status_code=404)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[str(c.value) if c.value is not None else "" for c in row] for row in ws.iter_rows()]
            if not rows:
                continue
            sheets.append({"sheet": name, "headers": rows[0], "rows": rows[1:]})
        wb.close()
        return JSONResponse(sheets)
    except Exception as e:
        logger.error(f"file-xlsx error for {safe_name}: {e}")
        return JSONResponse({"error": "Impossible de lire le fichier Excel."}, status_code=500)


@admin_router.delete("/api/admin/delete")
async def admin_delete(request: Request):
    require_monitoring(request)
    body = await request.json()
    filename = (body or {}).get("filename", "").strip()

    # Path traversal protection
    if not filename or any(s in filename for s in ("/", "\\", "..")):
        return JSONResponse({"error": "Nom de fichier invalide"}, status_code=400)

    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return JSONResponse({"error": "Fichier introuvable"}, status_code=404)

    os.remove(path)

    # Suppression immédiate des chunks Qdrant liés au fichier pour garantir la cohérence.
    try:
        from src.ingestion.vector_store import get_store, remove_files
        # get_store et remove_files sont synchrones
        store = get_store(force_reindex=False)
        await asyncio.to_thread(remove_files, store, {filename})
    except Exception as e:
        logger.warning(f"Delete Qdrant immédiat échoué: {e}")

    reindex_warning = None
    changed = None
    try:
        changed = bool(await asyncio.to_thread(index_data, force_reindex=False))
    except Exception as e:
        reindex_warning = str(e)
        logger.warning(f"Delete réindexation automatique échouée: {e}")

    qdrant_points = _count_points_for_filename(filename)

    await track("reindex", detail=f"suppression : {filename}")
    logger.info(f"Fichier supprimé : {filename}")
    payload = {"message": f"'{filename}' supprimé et index mis à jour."}
    payload["ingestion"] = {
        "changed": changed,
        "qdrant_points": qdrant_points,
        "verified": (qdrant_points == 0) if qdrant_points is not None else False,
        "summary": (
            f"Ingestion {'mise à jour' if changed else 'sans changement détecté'}; "
            + (f"Qdrant={qdrant_points} chunk(s) restant(s)." if qdrant_points is not None else "Qdrant non vérifiable.")
        ),
    }
    if reindex_warning:
        payload["warning"] = "Fichier supprimé, mais la réindexation automatique a échoué."
    return payload


# ═══════════════════════════════════════════════════════════════════════════════
# TENANTS — CRUD Admin (B2B)
# ═══════════════════════════════════════════════════════════════════════════════

import hashlib
import secrets
from datetime import datetime, timezone


async def _ensure_tenants_table():
    """Crée les tables B2B si elles n'existent pas (idempotent)."""
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return
        schema_path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "docs", "supabase_schema.sql")
        if os.path.exists(schema_path):
            with open(schema_path) as f:
                sql = f.read()
            for stmt in sql.split(";"):
                stmt = stmt.strip()
                if stmt and not stmt.startswith("--"):
                    try:
                        await supa.rpc("exec_sql", {"sql": stmt})
                    except Exception:
                        pass
    except Exception:
        pass


async def _get_tenant_for_user(user_id: str) -> dict | None:
    """Récupère le tenant associé à un utilisateur."""
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return None
        res = await supa.table("user_roles").select("tenant_id, role").eq("user_id", user_id).limit(1).execute()
        if not res.data:
            return None
        tenant_res = await supa.table("tenants").select("*").eq("id", res.data[0]["tenant_id"]).limit(1).execute()
        return tenant_res.data[0] if tenant_res.data else None
    except Exception:
        return None


@admin_router.get("/api/admin/tenants")
async def list_tenants(request: Request):
    """Liste tous les tenants (admin)."""
    require_monitoring(request)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        res = await supa.table("tenants").select("*").order("created_at", desc=True).execute()
        tenants = []
        for t in (res.data or []):
            members_res = await supa.table("user_roles").select("id", count="exact").eq("tenant_id", t["id"]).execute()
            owner_email = ""
            try:
                u = await supa.auth.admin.get_user_by_id(t["owner_id"])
                owner_email = u.user.email if u and u.user else ""
            except Exception:
                pass
            tenants.append({
                "id": t["id"], "name": t["name"], "slug": t.get("slug", ""),
                "plan": t.get("plan", "free"), "is_active": t.get("is_active", True),
                "owner_email": owner_email,
                "members": members_res.count if hasattr(members_res, 'count') else len(members_res.data or []),
                "created_at": str(t.get("created_at", "")),
            })
        return {"tenants": tenants, "total": len(tenants)}
    except Exception as e:
        logger.error(f"[TENANTS] Erreur liste: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.post("/api/admin/tenants")
async def create_tenant(request: Request):
    """Crée un nouveau tenant B2B."""
    require_monitoring(request)
    body = await request.json()
    name = (body or {}).get("name", "").strip()
    slug = (body or {}).get("slug", "").strip().lower().replace(" ", "-")
    owner_email = (body or {}).get("owner_email", "").strip().lower()
    plan = (body or {}).get("plan", "free").strip().lower()
    if not name or not slug or not owner_email:
        return JSONResponse({"error": "name, slug, et owner_email requis."}, status_code=400)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        existing = await supa.table("tenants").select("id").eq("slug", slug).execute()
        if existing.data:
            return JSONResponse({"error": f"Slug '{slug}' déjà utilisé."}, status_code=409)
        owner_id = None
        try:
            res = await supa.auth.admin.invite_user_by_email(owner_email)
            owner_id = str(res.user.id) if res and res.user else None
        except Exception as e:
            if "already registered" in str(e).lower():
                users = await supa.auth.admin.list_users()
                raw = users.users if hasattr(users, 'users') else (users if isinstance(users, list) else [])
                for u in raw:
                    if getattr(u, 'email', '').lower() == owner_email:
                        owner_id = str(u.id)
                        break
        if not owner_id:
            return JSONResponse({"error": "Impossible de créer/trouver l'utilisateur."}, status_code=500)
        tenant_res = await supa.table("tenants").insert({
            "name": name, "slug": slug, "owner_id": owner_id, "plan": plan,
            "max_users": 10 if plan == "pro" else (50 if plan == "enterprise" else 5),
        }).execute()
        tenant = tenant_res.data[0] if tenant_res.data else None
        if not tenant:
            return JSONResponse({"error": "Échec création."}, status_code=500)
        await supa.table("user_roles").insert({
            "user_id": owner_id, "tenant_id": tenant["id"], "role": "owner",
        }).execute()
        await track("tenant_create", detail=f"slug={slug} name={name}")
        return {"message": f"Tenant '{name}' créé.", "tenant": {"id": tenant["id"], "name": name, "slug": slug, "plan": plan}}
    except Exception as e:
        logger.error(f"[TENANTS] Erreur création: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.delete("/api/admin/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, request: Request):
    """Supprime un tenant et ses données."""
    require_monitoring(request)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        try:
            from src.ingestion.vector_store import _get_client as _qdrant, _COLLECTION_NAME
            from qdrant_client.models import Filter, FieldCondition, MatchValue
            qdrant = _qdrant()
            qdrant.delete(collection_name=_COLLECTION_NAME, points_selector=Filter(
                must=[FieldCondition(key="metadata.tenant_id", match=MatchValue(value=tenant_id))]))
        except Exception:
            pass
        import shutil
        tenant_dir = os.path.join(DATA_DIR, tenant_id)
        if os.path.exists(tenant_dir):
            shutil.rmtree(tenant_dir, ignore_errors=True)
        await supa.table("tenants").delete().eq("id", tenant_id).execute()
        await track("tenant_delete", detail=f"tenant_id={tenant_id}")
        return {"message": "Tenant supprimé.", "tenant_id": tenant_id}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.get("/api/admin/metrics")
async def admin_metrics(request: Request):
    """Métriques globales + par tenant (admin)."""
    require_monitoring(request)
    try:
        from src.monitoring.tracker import _get_client
        from datetime import timedelta
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        total_reqs = await supa.table("usage_metrics").select("requests", "tokens_input", "tokens_output").execute()
        total_requests = sum(r.get("requests", 0) for r in (total_reqs.data or []))
        total_tokens_in = sum(r.get("tokens_input", 0) for r in (total_reqs.data or []))
        total_tokens_out = sum(r.get("tokens_output", 0) for r in (total_reqs.data or []))
        week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
        tenants_res = await supa.table("tenants").select("id, name, slug, plan").execute()
        tenants_usage = []
        for t in (tenants_res.data or []):
            usage_res = await supa.table("usage_metrics").select("*").eq("tenant_id", t["id"]).gte("date", week_ago).execute()
            week_reqs = sum(r.get("requests", 0) for r in (usage_res.data or []))
            week_tokens = sum(r.get("tokens_input", 0) + r.get("tokens_output", 0) for r in (usage_res.data or []))
            tenants_usage.append({"tenant_id": t["id"], "name": t["name"], "slug": t["slug"], "plan": t["plan"],
                                  "week_requests": week_reqs, "week_tokens": week_tokens})
        return {"global": {"total_requests": total_requests, "total_tokens_input": total_tokens_in, "total_tokens_output": total_tokens_out},
                "tenants": sorted(tenants_usage, key=lambda x: x["week_tokens"], reverse=True)}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


def _generate_api_key() -> tuple[str, str, str]:
    raw = "rk_" + secrets.token_urlsafe(32)
    return raw, hashlib.sha256(raw.encode()).hexdigest(), raw[:10]


@admin_router.get("/api/tenant/api-keys")
async def tenant_api_keys(user_id: str = Depends(get_current_user)):
    tenant = await _get_tenant_for_user(user_id)
    if not tenant:
        return JSONResponse({"error": "Aucun tenant associé."}, status_code=404)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        res = await supa.table("api_keys").select("id, name, key_prefix, is_active, last_used, created_at").eq("tenant_id", tenant["id"]).execute()
        return {"keys": res.data or [], "tenant_id": tenant["id"]}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.post("/api/tenant/api-keys")
async def create_api_key(request: Request, user_id: str = Depends(get_current_user)):
    tenant = await _get_tenant_for_user(user_id)
    if not tenant:
        return JSONResponse({"error": "Aucun tenant associé."}, status_code=404)
    body = await request.json()
    name = (body or {}).get("name", "Default").strip() or "Default"
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        raw_key, key_hash, prefix = _generate_api_key()
        await supa.table("api_keys").insert({"tenant_id": tenant["id"], "name": name, "key_hash": key_hash, "key_prefix": prefix}).execute()
        await track("api_key_create", detail=f"tenant={tenant['id']}")
        return {"message": "Clé API créée. Conservez-la — plus jamais affichée.", "api_key": raw_key, "prefix": prefix, "name": name}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.delete("/api/tenant/api-keys/{key_id}")
async def revoke_api_key(key_id: str, user_id: str = Depends(get_current_user)):
    tenant = await _get_tenant_for_user(user_id)
    if not tenant:
        return JSONResponse({"error": "Aucun tenant associé."}, status_code=404)
    try:
        from src.monitoring.tracker import _get_client
        supa = await _get_client()
        if not supa:
            return JSONResponse({"error": "Supabase non configuré."}, status_code=503)
        existing = await supa.table("api_keys").select("id").eq("id", key_id).eq("tenant_id", tenant["id"]).execute()
        if not existing.data:
            return JSONResponse({"error": "Clé introuvable."}, status_code=404)
        await supa.table("api_keys").update({"is_active": False}).eq("id", key_id).execute()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@admin_router.get("/api/tenant/me")
async def tenant_me(user_id: str = Depends(get_current_user)):
    tenant = await _get_tenant_for_user(user_id)
    if not tenant:
        return {"tenant": None, "mode": "solo", "message": "Pas de tenant B2B. Mode individuel."}
    return {"tenant": tenant, "mode": "b2b"}


# ── Usage tracking (appelé depuis routes.py) ──────────────────────────────────

async def _increment_usage(user_id: str, tokens_in: int = 0, tokens_out: int = 0):
    """Incrémente les métriques d'usage pour le tenant de l'utilisateur (background)."""
    try:
        from src.monitoring.tracker import _get_client
        from datetime import date
        supa = await _get_client()
        if not supa:
            return
        tenant = await _get_tenant_for_user(user_id)
        if not tenant:
            return
        today = date.today().isoformat()
        existing = await supa.table("usage_metrics").select("id, requests, tokens_input, tokens_output") \
            .eq("tenant_id", tenant["id"]).eq("date", today).execute()
        if existing.data:
            row = existing.data[0]
            await supa.table("usage_metrics").update({
                "requests": row["requests"] + 1,
                "tokens_input": row["tokens_input"] + tokens_in,
                "tokens_output": row["tokens_output"] + tokens_out,
            }).eq("id", row["id"]).execute()
        else:
            await supa.table("usage_metrics").insert({
                "tenant_id": tenant["id"],
                "date": today,
                "requests": 1,
                "tokens_input": tokens_in,
                "tokens_output": tokens_out,
            }).execute()
    except Exception:
        pass  # Fail-open — ne bloque jamais la réponse
