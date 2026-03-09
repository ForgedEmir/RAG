"""
Validation et filtrage des fichiers en entrée.
Ce module vérifie que tous les fichiers respectent les critères de qualité avant ingestion :
- Encodage UTF-8 valide
- Taille de fichier acceptable (ni trop petit, ni trop gros)
- Format CSV avec colonnes valides
- Structure JSON conforme
- Contenu non vide ou significatif
"""
import os
import csv
import json
import logging
from typing import Optional, List, Dict, Tuple
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# Contraintes de taille
MIN_FILE_SIZE_BYTES = 10  # 10 octets minimum
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB maximum
MIN_MEANINGFUL_CONTENT_CHARS = 5  # Nombre minimum de caractères significatifs (hors espaces)

# Encodages autorisés (priorité UTF-8)
ALLOWED_ENCODINGS = ['utf-8', 'utf-8-sig']

# CSV : colonnes attendues (au moins une de ces colonnes doit être présente)
EXPECTED_CSV_COLUMNS = {
    # Colonnes fréquentes pour des données de lore
    'personnages': ['nom', 'name', 'description', 'role', 'faction'],
    'lieux': ['nom', 'name', 'description', 'location', 'lieu'],
    'objets': ['nom', 'name', 'description', 'type', 'item'],
    'evenements': ['nom', 'name', 'description', 'date', 'event', 'événement'],
    'factions': ['nom', 'name', 'description', 'chef', 'leader'],
}

# JSON : clés attendues au niveau racine (au moins une doit être présente pour être valide)
EXPECTED_JSON_ROOT_KEYS = ['nom', 'name', 'description', 'titre', 'title', 'data', 'items', 'id']


@dataclass
class ValidationResult:
    """Résultat d'une validation de fichier."""
    is_valid: bool
    filepath: str
    errors: List[str]
    warnings: List[str]
    
    def __repr__(self):
        status = "✓ VALIDE" if self.is_valid else "✗ INVALIDE"
        return f"{status} - {os.path.basename(self.filepath)}"


def validate_file(filepath: str) -> ValidationResult:
    """
    Point d'entrée principal pour valider un fichier.
    Effectue toutes les vérifications nécessaires selon le type de fichier.
    
    Args:
        filepath: Chemin absolu vers le fichier à valider
        
    Returns:
        ValidationResult avec le statut de validation et les erreurs/warnings
    """
    errors = []
    warnings = []
    
    # 1. Vérifier que le fichier existe
    if not os.path.exists(filepath):
        errors.append("Le fichier n'existe pas")
        return ValidationResult(False, filepath, errors, warnings)
    
    if not os.path.isfile(filepath):
        errors.append("Le chemin ne pointe pas vers un fichier")
        return ValidationResult(False, filepath, errors, warnings)
    
    filename = os.path.basename(filepath)
    
    # 2. Vérifier la taille du fichier
    file_size = os.path.getsize(filepath)
    
    if file_size < MIN_FILE_SIZE_BYTES:
        errors.append(f"Fichier trop petit ({file_size} octets, minimum {MIN_FILE_SIZE_BYTES} octets)")
        return ValidationResult(False, filepath, errors, warnings)
    
    if file_size > MAX_FILE_SIZE_BYTES:
        errors.append(f"Fichier trop volumineux ({file_size / (1024*1024):.2f} MB, maximum {MAX_FILE_SIZE_BYTES / (1024*1024):.0f} MB)")
        return ValidationResult(False, filepath, errors, warnings)
    
    if file_size < 100:
        warnings.append(f"Fichier très petit ({file_size} octets), vérifiez qu'il contient bien des données")
    
    # 3. Vérifier l'encodage UTF-8
    encoding_valid, encoding_error = _validate_encoding(filepath)
    if not encoding_valid:
        errors.append(f"Encodage invalide : {encoding_error}")
        return ValidationResult(False, filepath, errors, warnings)
    
    # 4. Vérifier le contenu selon l'extension
    _, extension = os.path.splitext(filename)
    extension = extension.lower()
    
    try:
        if extension == '.csv':
            csv_errors, csv_warnings = _validate_csv(filepath)
            errors.extend(csv_errors)
            warnings.extend(csv_warnings)
        
        elif extension == '.json':
            json_errors, json_warnings = _validate_json(filepath)
            errors.extend(json_errors)
            warnings.extend(json_warnings)
        
        elif extension in ['.txt', '.md']:
            text_errors, text_warnings = _validate_text_content(filepath)
            errors.extend(text_errors)
            warnings.extend(text_warnings)
        
        elif extension == '.xlsx':
            excel_errors, excel_warnings = _validate_excel(filepath)
            errors.extend(excel_errors)
            warnings.extend(excel_warnings)
        
        # Les autres extensions (.xml par exemple) passent la validation de base
        # mais ne font pas l'objet de vérifications spécifiques
        
    except Exception as e:
        errors.append(f"Erreur lors de la validation du contenu : {str(e)}")
        logger.exception(f"Exception lors de la validation de {filename}")
    
    # 5. Bilan final
    is_valid = len(errors) == 0
    
    if is_valid and warnings:
        logger.warning(f"Fichier validé avec avertissements : {filename}")
        for warning in warnings:
            logger.warning(f"  ⚠ {warning}")
    elif not is_valid:
        logger.error(f"Fichier rejeté : {filename}")
        for error in errors:
            logger.error(f"  ✗ {error}")
    
    return ValidationResult(is_valid, filepath, errors, warnings)


def _validate_encoding(filepath: str) -> Tuple[bool, Optional[str]]:
    """
    Vérifie que le fichier est bien encodé en UTF-8.
    
    Returns:
        (True, None) si l'encodage est valide
        (False, message_erreur) sinon
    """
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            # On essaie de lire tout le fichier
            _ = f.read()
        return True, None
    except UnicodeDecodeError as e:
        # Essayer de détecter l'encodage réel
        try:
            with open(filepath, 'rb') as f:
                raw_data = f.read(1024)  # Lire les premiers 1024 octets
            
            # Essayer quelques encodages courants
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    raw_data.decode(encoding)
                    return False, f"Le fichier semble être encodé en {encoding.upper()}, pas en UTF-8. Veuillez le convertir."
                except:
                    continue
            
            return False, f"Erreur de décodage UTF-8 : {str(e)}"
        except Exception:
            return False, "Impossible de lire le fichier"
    except Exception as e:
        return False, f"Erreur lors de la lecture : {str(e)}"


def _validate_text_content(filepath: str) -> Tuple[List[str], List[str]]:
    """
    Valide le contenu des fichiers texte (.txt, .md).
    Vérifie que le fichier n'est pas vide ou ne contient que des espaces.
    """
    errors = []
    warnings = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Vérifier que le contenu n'est pas vide
        if not content.strip():
            errors.append("Le fichier est vide ou ne contient que des espaces")
            return errors, warnings
        
        # Vérifier qu'il y a un minimum de contenu significatif
        meaningful_chars = len([c for c in content if not c.isspace()])
        if meaningful_chars < MIN_MEANINGFUL_CONTENT_CHARS:
            errors.append(f"Contenu insuffisant : {meaningful_chars} caractères significatifs (minimum {MIN_MEANINGFUL_CONTENT_CHARS})")
        
        # Warning si le contenu est très court
        if 5 <= meaningful_chars < 20:
            warnings.append(f"Contenu très court : seulement {meaningful_chars} caractères")
        
    except Exception as e:
        errors.append(f"Erreur lors de la lecture du contenu : {str(e)}")
    
    return errors, warnings


def _validate_csv(filepath: str) -> Tuple[List[str], List[str]]:
    """
    Valide un fichier CSV :
    - Doit pouvoir être parsé correctement
    - Doit avoir au moins une ligne de données (en plus de l'en-tête)
    - Les colonnes doivent correspondre à un type de données reconnu
    """
    errors = []
    warnings = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        if not content.strip():
            errors.append("Le fichier CSV est vide")
            return errors, warnings
        
        # Essayer de détecter le dialecte
        try:
            dialect = csv.Sniffer().sniff(content[:1024])
        except csv.Error:
            # Si le sniffer échoue, utiliser le dialecte par défaut
            dialect = csv.excel
            warnings.append("Format CSV non standard détecté, utilisation du format par défaut")
        
        # Parser le CSV
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        if not rows:
            errors.append("Le fichier CSV ne contient aucune ligne")
            return errors, warnings
        
        # Vérifier l'en-tête
        if len(rows) < 2:
            errors.append("Le fichier CSV ne contient pas de données (seulement l'en-tête ou moins)")
            return errors, warnings
        
        header = [col.strip().lower() for col in rows[0]]
        
        if not header or all(not col for col in header):
            errors.append("L'en-tête du CSV est vide ou invalide")
            return errors, warnings
        
        # Vérifier que les colonnes correspondent à un type connu
        recognized = False
        for csv_type, expected_cols in EXPECTED_CSV_COLUMNS.items():
            if any(col in header for col in expected_cols):
                recognized = True
                logger.info(f"CSV reconnu comme type '{csv_type}' avec colonnes : {', '.join(header)}")
                break
        
        if not recognized:
            all_expected = []
            for cols in EXPECTED_CSV_COLUMNS.values():
                all_expected.extend(cols)
            unique_expected = sorted(set(all_expected))
            warnings.append(
                f"Colonnes CSV non reconnues : {', '.join(header)}. "
                f"Colonnes attendues (au moins une) : {', '.join(unique_expected)}"
            )
        
        # Vérifier qu'il y a des données non vides
        data_rows = rows[1:]
        non_empty_rows = [row for row in data_rows if any(cell.strip() for cell in row)]
        
        if not non_empty_rows:
            errors.append("Le fichier CSV ne contient aucune donnée valide après l'en-tête")
        elif len(non_empty_rows) < len(data_rows):
            warnings.append(f"{len(data_rows) - len(non_empty_rows)} ligne(s) vide(s) détectée(s)")
        
        # Vérifier la cohérence du nombre de colonnes
        expected_col_count = len(header)
        inconsistent_rows = [i+2 for i, row in enumerate(data_rows) if len(row) != expected_col_count]
        if inconsistent_rows:
            if len(inconsistent_rows) <= 5:
                warnings.append(f"Lignes avec nombre de colonnes incohérent : {', '.join(map(str, inconsistent_rows))}")
            else:
                warnings.append(f"{len(inconsistent_rows)} ligne(s) avec nombre de colonnes incohérent")
    
    except Exception as e:
        errors.append(f"Impossible de parser le CSV : {str(e)}")
    
    return errors, warnings


def _validate_json(filepath: str) -> Tuple[List[str], List[str]]:
    """
    Valide un fichier JSON :
    - Doit être un JSON valide syntaxiquement
    - Doit contenir des données (pas juste {} ou [])
    - Doit avoir une structure reconnue (clés attendues)
    """
    errors = []
    warnings = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        if not content.strip():
            errors.append("Le fichier JSON est vide")
            return errors, warnings
        
        # Parser le JSON
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            errors.append(f"JSON invalide : {str(e)}")
            return errors, warnings
        
        # Vérifier que le JSON n'est pas vide
        if data is None:
            errors.append("Le JSON contient 'null'")
            return errors, warnings
        
        if isinstance(data, dict) and not data:
            errors.append("Le JSON est un objet vide {}")
            return errors, warnings
        
        if isinstance(data, list) and not data:
            errors.append("Le JSON est un tableau vide []")
            return errors, warnings
        
        # Vérifier la structure
        recognized = False
        
        if isinstance(data, dict):
            keys = [k.lower() for k in data.keys()]
            if any(expected_key in keys for expected_key in EXPECTED_JSON_ROOT_KEYS):
                recognized = True
                logger.info(f"JSON reconnu avec clés : {', '.join(data.keys())}")
        
        elif isinstance(data, list) and len(data) > 0:
            # Si c'est un tableau, vérifier le premier élément
            first_item = data[0]
            if isinstance(first_item, dict):
                keys = [k.lower() for k in first_item.keys()]
                if any(expected_key in keys for expected_key in EXPECTED_JSON_ROOT_KEYS):
                    recognized = True
                    logger.info(f"JSON tableau reconnu avec {len(data)} élément(s)")
        
        if not recognized:
            warnings.append(
                f"Structure JSON non reconnue. Clés attendues (au moins une) : {', '.join(EXPECTED_JSON_ROOT_KEYS)}"
            )
        
        # Vérifier la profondeur (warning si trop profond)
        depth = _get_json_depth(data)
        if depth > 10:
            warnings.append(f"Structure JSON très profonde ({depth} niveaux), cela peut affecter la qualité de l'indexation")
    
    except Exception as e:
        errors.append(f"Erreur lors de la validation JSON : {str(e)}")
    
    return errors, warnings


def _validate_excel(filepath: str) -> Tuple[List[str], List[str]]:
    """
    Valide un fichier Excel :
    - Doit pouvoir être ouvert
    - Doit contenir au moins une feuille avec des données
    """
    errors = []
    warnings = []
    
    try:
        import openpyxl
        
        # Ouvrir le fichier
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
        except Exception as e:
            errors.append(f"Impossible d'ouvrir le fichier Excel : {str(e)}")
            return errors, warnings
        
        # Vérifier qu'il y a au moins une feuille
        if not workbook.sheetnames:
            errors.append("Le fichier Excel ne contient aucune feuille")
            workbook.close()
            return errors, warnings
        
        # Vérifier qu'au moins une feuille contient des données
        has_data = False
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.rows)
            
            if len(rows) >= 2:  # Au moins en-tête + 1 ligne de données
                # Vérifier qu'il y a des cellules non vides
                for row in rows:
                    if any(cell.value is not None and str(cell.value).strip() for cell in row):
                        has_data = True
                        break
            
            if has_data:
                break
        
        if not has_data:
            errors.append("Le fichier Excel ne contient aucune donnée valide")
        
        workbook.close()
    
    except ImportError:
        warnings.append("Module openpyxl non disponible, validation Excel limitée")
    except Exception as e:
        errors.append(f"Erreur lors de la validation Excel : {str(e)}")
    
    return errors, warnings


def _get_json_depth(obj, current_depth=0):
    """Calcule récursivement la profondeur maximale d'une structure JSON."""
    if isinstance(obj, dict):
        if not obj:
            return current_depth
        return max(_get_json_depth(v, current_depth + 1) for v in obj.values())
    elif isinstance(obj, list):
        if not obj:
            return current_depth
        return max(_get_json_depth(item, current_depth + 1) for item in obj)
    else:
        return current_depth


def validate_files_batch(filepaths: List[str]) -> Dict[str, ValidationResult]:
    """
    Valide un lot de fichiers et retourne les résultats.
    
    Args:
        filepaths: Liste des chemins de fichiers à valider
        
    Returns:
        Dictionnaire {filepath: ValidationResult}
    """
    results = {}
    
    logger.info(f"Début de la validation de {len(filepaths)} fichier(s)...")
    
    for filepath in filepaths:
        result = validate_file(filepath)
        results[filepath] = result
    
    # Statistiques
    valid_count = sum(1 for r in results.values() if r.is_valid)
    invalid_count = len(results) - valid_count
    
    logger.info(f"Validation terminée : {valid_count} valide(s), {invalid_count} rejeté(s)")
    
    return results
